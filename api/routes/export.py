"""
エクスポートAPIエンドポイント
Excel形式でのデータエクスポート機能を提供する。

GET /api/export/asw-matrix
  → 5シート構成の Excel をストリーム返却
  Sheet 1: ASW_SASA_5Y   – 残存5年以下クーポン国債の SASA ASW [bps]
  Sheet 2: TONA_HIST_1Y  – TONA OIS スワップ金利履歴 [%]
  Sheet 3: TONA_FWD_3M   – キャリーロール: 3M×(N-3M) fwd − Spot(N) [bps]
  Sheet 4: TONA_INST_FWD – 瞬間フォワード: (N-3M)×3M fwd レート [%]
  Sheet 5: TONA_IMM      – IMMストリップ: 各IMMロール3M OIS fwd [%]

レイアウト (全シート共通):
  行1       : 列ヘッダー（bond_code / tenor / IMMコード）
  行2       : 参照行（due_date / 単位 / 実IMM日付）
  行3-7     : Z-score (10D/20D/50D/100D/200D) Excel数式
  行8-12    : 空白（ユーザー用スペース）
  行13以降  : データ（新しい日付から降順）
  固定表示  : 行1-12 + 列A (freeze_panes = B13)

ヒートマップ:
  Z-score 行3-7  : 行単位 red→white→blue
  データ 行13-32 : 列単位 red→white→blue
"""

import io
import logging
from datetime import date, timedelta
from typing import Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from starlette.concurrency import run_in_threadpool
from zoneinfo import ZoneInfo

from core.db.engine import AsyncSessionLocal
from core.calculations.bond_math import QuantLibHelper, _ql_lock
from core.utils.date_utils import third_wednesday as _third_wednesday_core, get_imm_strip_columns as _get_imm_strip_columns_core

_JST = ZoneInfo("Asia/Tokyo")


def _jst_today() -> date:
    """Cloud Run (UTC) 環境で日本時間の「今日」を返す"""
    from datetime import datetime
    return datetime.now(_JST).date()


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

router = APIRouter()
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# 定数
# ─────────────────────────────────────────────
DATA_START_ROW = 13
MAX_MATURITY_YEARS = 5
DEFAULT_LOOKBACK_DAYS = 365

ZSCORE_ROWS = [
    (3,  "Z 10D",   10),
    (4,  "Z 20D",   20),
    (5,  "Z 50D",   50),
    (6,  "Z 100D", 100),
    (7,  "Z 200D", 200),
]

STANDARD_OIS_TENORS = [
    "2W", "1M", "2M", "3M", "4M", "5M", "6M", "7M", "8M", "9M", "11M",
    "1Y", "15M", "18M",
    "2Y", "3Y", "4Y", "5Y", "6Y", "7Y", "8Y", "9Y", "10Y",
    "11Y", "12Y", "15Y", "20Y", "25Y", "30Y", "35Y", "40Y",
]

# ─────────────────────────────────────────────
# Excel スタイル
# ─────────────────────────────────────────────
_H1_FILL   = PatternFill("solid", fgColor="1F4E79")
_H1_FONT   = Font(color="FFFFFF", bold=True, size=9)
_H2_FILL   = PatternFill("solid", fgColor="2E75B6")
_H2_FONT   = Font(color="FFFFFF", bold=True, size=8)
_ZS_FILL   = PatternFill("solid", fgColor="FFF2CC")
_ZS_FONT   = Font(color="7F6000", bold=True, size=8)
_DATE_FILL = PatternFill("solid", fgColor="D6E4F0")
_DATE_FONT = Font(bold=True, size=9)
_DATA_FONT = Font(size=9)
_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT  = Alignment(horizontal="right",  vertical="center")


# ─────────────────────────────────────────────
# ユーティリティ
# ─────────────────────────────────────────────
def _tenor_to_months(tenor: str) -> float:
    t = tenor.upper()
    if t.endswith("Y"):
        return float(t[:-1]) * 12
    if t.endswith("M"):
        return float(t[:-1])
    if t.endswith("W"):
        return float(t[:-1]) / 4.33
    if t.endswith("D"):
        return float(t[:-1]) / 30.4
    return 0.0


def _third_wednesday(year: int, month: int) -> date:
    return _third_wednesday_core(year, month)


def _get_imm_strip_columns(from_date: date, years: int = 10, max_cols: int = 40) -> list:
    return _get_imm_strip_columns_core(from_date, years=years, max_cols=max_cols)


def _cell(ws, row: int, col: int, value, fill, font, alignment):
    c = ws.cell(row, col, value)
    c.fill, c.font, c.alignment, c.border = fill, font, alignment, _THIN
    return c


def _write_header_rows(ws, col_labels: list, row2_labels: list):
    _cell(ws, 1, 1, "trade_date", _H1_FILL, _H1_FONT, _CENTER)
    for ci, label in enumerate(col_labels, 2):
        _cell(ws, 1, ci, label, _H1_FILL, _H1_FONT, _CENTER)
    _cell(ws, 2, 1, row2_labels[0] if row2_labels else "", _H2_FILL, _H2_FONT, _CENTER)
    for ci, label in enumerate(row2_labels[1:] if row2_labels else [], 2):
        _cell(ws, 2, ci, label, _H2_FILL, _H2_FONT, _CENTER)


def _write_zscore_rows(ws, col_count: int):
    for row_idx, label, window in ZSCORE_ROWS:
        end_row = DATA_START_ROW + window - 1
        _cell(ws, row_idx, 1, label, _ZS_FILL, _ZS_FONT, _CENTER)
        for ci in range(2, col_count + 2):
            cl = get_column_letter(ci)
            formula = (
                f'=IFERROR(({cl}{DATA_START_ROW}-AVERAGE({cl}{DATA_START_ROW}:{cl}{end_row}))'
                f'/STDEV({cl}{DATA_START_ROW}:{cl}{end_row}),"")'
            )
            c = ws.cell(row_idx, ci, formula)
            c.fill, c.font, c.alignment, c.border = _ZS_FILL, _ZS_FONT, _RIGHT, _THIN
            c.number_format = "0.00"


def _write_data_rows(ws, matrix_rows: list, col_keys: list, n_format: str = "0.00"):
    for ri, (trade_date, row_vals) in enumerate(matrix_rows, DATA_START_ROW):
        _cell(ws, ri, 1, str(trade_date), _DATE_FILL, _DATE_FONT, _CENTER)
        for ci, key in enumerate(col_keys, 2):
            val = row_vals.get(key)
            c = ws.cell(ri, ci)
            if val is not None:
                c.value = float(val)
                c.number_format = n_format
            c.font, c.alignment, c.border = _DATA_FONT, _RIGHT, _THIN


# ─────────────────────────────────────────────
# ヒートマップ
# ─────────────────────────────────────────────
def _apply_heatmap(ws, min_col: int, max_col: int, min_row: int, max_row: int):
    """赤(小)→白(中)→青(大) のカラースケールを適用"""
    rule = ColorScaleRule(
        start_type='min',        start_color='FF0000',
        mid_type='percentile',   mid_value=50, mid_color='FFFFFF',
        end_type='max',          end_color='0070C0',
    )
    rng = (
        f"{get_column_letter(min_col)}{min_row}"
        f":{get_column_letter(max_col)}{max_row}"
    )
    ws.conditional_formatting.add(rng, rule)


def _apply_zscore_heatmap(ws, n_data_cols: int):
    """Z-score 5行×n列のブロック全体に1つのカラースケールを適用"""
    if n_data_cols <= 0:
        return
    first_row = ZSCORE_ROWS[0][0]   # 3
    last_row  = ZSCORE_ROWS[-1][0]  # 7
    _apply_heatmap(ws, 2, n_data_cols + 1, first_row, last_row)


def _apply_data_heatmap(ws, n_data_cols: int):
    """データ先頭20行(13-32)に列ごとのヒートマップを適用"""
    if n_data_cols <= 0:
        return
    data_end_row = DATA_START_ROW + 19
    for ci in range(2, n_data_cols + 2):
        _apply_heatmap(ws, ci, ci, DATA_START_ROW, data_end_row)


def _set_sheet_style(ws, col_count: int, col_a_width: float = 12, data_col_width: float = 11):
    ws.column_dimensions["A"].width = col_a_width
    for ci in range(2, col_count + 2):
        ws.column_dimensions[get_column_letter(ci)].width = data_col_width
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16
    for row_idx, _, _ in ZSCORE_ROWS:
        ws.row_dimensions[row_idx].height = 15
    ws.freeze_panes = "B13"
    _apply_zscore_heatmap(ws, col_count)
    _apply_data_heatmap(ws, col_count)


# ─────────────────────────────────────────────
# シート構築
# ─────────────────────────────────────────────
def _build_asw_sheet(wb: Workbook, bond_codes: list, due_dates: dict,
                     matrix_rows: list, sheet_name: str):
    ws = wb.create_sheet(sheet_name)
    row2 = ["due_date"] + [due_dates.get(c, "") for c in bond_codes]
    _write_header_rows(ws, bond_codes, row2)
    _write_zscore_rows(ws, len(bond_codes))
    _write_data_rows(ws, matrix_rows, bond_codes, "0.00")
    _set_sheet_style(ws, len(bond_codes))


def _build_tona_hist_sheet(wb: Workbook, tenors: list,
                            matrix_rows: list, sheet_name: str):
    ws = wb.create_sheet(sheet_name)
    row2 = ["%"] + ["%"] * len(tenors)
    _write_header_rows(ws, tenors, row2)
    _write_zscore_rows(ws, len(tenors))
    _write_data_rows(ws, matrix_rows, tenors, "0.000")
    _set_sheet_style(ws, len(tenors))


def _build_carry_roll_sheet(wb: Workbook, tenors: list,
                             matrix_rows: list, sheet_name: str):
    """Sheet3: 3M×(N-3M) fwd − Spot(N) [bps]"""
    ws = wb.create_sheet(sheet_name)
    row2 = ["bps"] + ["bps"] * len(tenors)
    _write_header_rows(ws, tenors, row2)
    _write_zscore_rows(ws, len(tenors))
    _write_data_rows(ws, matrix_rows, tenors, "0.00")
    _set_sheet_style(ws, len(tenors))


def _build_tona_inst_fwd_sheet(wb: Workbook, tenors: list,
                                matrix_rows: list, sheet_name: str):
    """Sheet4: (N-3M)×3M 瞬間フォワードレート [%]"""
    ws = wb.create_sheet(sheet_name)
    row2 = ["%"] + ["%"] * len(tenors)
    _write_header_rows(ws, tenors, row2)
    _write_zscore_rows(ws, len(tenors))
    _write_data_rows(ws, matrix_rows, tenors, "0.000")
    _set_sheet_style(ws, len(tenors))


def _build_imm_strip_sheet(wb: Workbook, imm_cols: list,
                            matrix_rows: list, sheet_name: str):
    """Sheet5: IMM日付開始の3M OIS fwdレート [%]
    imm_cols: [(code, date_str), ...]  e.g. [("U26", "2026-09-16"), ...]
    """
    ws = wb.create_sheet(sheet_name)
    col_labels  = [code for code, _ in imm_cols]
    row2_labels = ["date"] + [ds for _, ds in imm_cols]
    _write_header_rows(ws, col_labels, row2_labels)
    _write_zscore_rows(ws, len(imm_cols))
    _write_data_rows(ws, matrix_rows, col_labels, "0.000")
    _set_sheet_style(ws, len(imm_cols), data_col_width=9)


_DAYS_FILL = PatternFill("solid", fgColor="E2EFDA")   # 薄緑
_DAYS_FONT = Font(color="375623", bold=True, size=8)   # 濃緑

# IMM 単体 Excel 専用レイアウト定数
_IMM_DATA_START_ROW = 14   # 行3(日数)が追加された分だけ DATA_START_ROW より1行後ろ
_IMM_ZSCORE_ROWS = [
    (4,  "Z 10D",   10),
    (5,  "Z 20D",   20),
    (6,  "Z 50D",   50),
    (7,  "Z 100D", 100),
    (8,  "Z 200D", 200),
]


def _compute_imm_days(imm_cols: list, reference_date: date) -> list[int]:
    """reference_date（今日）から各IMM日付までの暦日数を返す。"""
    return [(date.fromisoformat(ds) - reference_date).days for _, ds in imm_cols]


def _build_imm_standalone_sheet(wb: Workbook, imm_cols: list,
                                  matrix_rows: list, sheet_name: str,
                                  reference_date: date = None):
    """IMM単体Excel用シート。

    行1 : IMMコード (U26, Z26, H27 ...)
    行2 : 実際のIMM日付 (第3水曜日の日付文字列)
    行3 : 今日からIMM日付までの暦日数
    行4-8: Z-score (10D/20D/50D/100D/200D) ← DATA_START_ROW=14 参照
    行9-13: 空白
    行14- : データ (trade_date 降順)

    imm_cols: [(code, date_str), ...]
    reference_date: 日数計算の基準日（今日）
    """
    ws = wb.create_sheet(sheet_name)
    col_labels  = [code for code, _ in imm_cols]
    row2_labels = ["IMM date"] + [ds for _, ds in imm_cols]
    n_cols = len(imm_cols)

    # ── 行1・行2: コード・日付ヘッダー ──────────────────────────
    _write_header_rows(ws, col_labels, row2_labels)

    # ── 行3: 今日からIMM日付までの暦日数 ──────────────────────
    ref = reference_date or _jst_today()
    _cell(ws, 3, 1, "days from today", _DAYS_FILL, _DAYS_FONT, _CENTER)
    days_list = _compute_imm_days(imm_cols, ref)
    for ci, d in enumerate(days_list, 2):
        c = ws.cell(3, ci, d)
        c.fill, c.font, c.alignment, c.border = _DAYS_FILL, _DAYS_FONT, _CENTER, _THIN
        c.number_format = "0"

    # ── 行4-8: Z-score ────────────────────────────────────────
    data_row = _IMM_DATA_START_ROW
    for row_idx, label, window in _IMM_ZSCORE_ROWS:
        end_row = data_row + window - 1
        _cell(ws, row_idx, 1, label, _ZS_FILL, _ZS_FONT, _CENTER)
        for ci in range(2, n_cols + 2):
            cl = get_column_letter(ci)
            formula = (
                f'=IFERROR(({cl}{data_row}-AVERAGE({cl}{data_row}:{cl}{end_row}))'
                f'/STDEV({cl}{data_row}:{cl}{end_row}),"")'
            )
            c = ws.cell(row_idx, ci, formula)
            c.fill, c.font, c.alignment, c.border = _ZS_FILL, _ZS_FONT, _RIGHT, _THIN
            c.number_format = "0.00"

    # ── 行14-: データ ─────────────────────────────────────────
    for ri, (trade_date, row_vals) in enumerate(matrix_rows, data_row):
        _cell(ws, ri, 1, str(trade_date), _DATE_FILL, _DATE_FONT, _CENTER)
        for ci, key in enumerate(col_labels, 2):
            val = row_vals.get(key)
            c = ws.cell(ri, ci)
            if val is not None:
                c.value = float(val)
                c.number_format = "0.0000"
            c.font, c.alignment, c.border = _DATA_FONT, _RIGHT, _THIN

    # ── スタイル・列幅・ペイン固定 ─────────────────────────────
    ws.column_dimensions["A"].width = 13
    for ci in range(2, n_cols + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 15
    for row_idx, _, _ in _IMM_ZSCORE_ROWS:
        ws.row_dimensions[row_idx].height = 15
    ws.freeze_panes = f"B{data_row}"

    # Z-score ヒートマップ (行4-8)
    first_zs = _IMM_ZSCORE_ROWS[0][0]
    last_zs  = _IMM_ZSCORE_ROWS[-1][0]
    _apply_heatmap(ws, 2, n_cols + 1, first_zs, last_zs)

    # データ先頭20行ヒートマップ (行14-33)
    data_end = data_row + 19
    for ci in range(2, n_cols + 2):
        _apply_heatmap(ws, ci, ci, data_row, data_end)


# ─────────────────────────────────────────────
# QuantLib 全フォワードレート計算（同期・スレッドプール用）
# ─────────────────────────────────────────────
def _compute_all_forward_rates_sync(
    dates_ois: dict,      # {date_str: {tenor: rate_pct}}
    carry_tenors: list,   # キャリーロール対象テナー (>3M)
    imm_cols: list,       # [(code, date_str), ...]
) -> tuple:               # (carry_results, inst_results, imm_results_3m, imm_results_1y)
    """
    全日付について OIS カーブを1回だけ構築し、以下を一括計算:
      carry_results   : {date_str: {tenor: bps}}  3M×(N-3M) fwd - spot
      inst_results    : {date_str: {tenor: %}}    (N-3M)×3M fwd
      imm_results_3m  : {date_str: {code: %}}     IMM日付開始 3M fwd
      imm_results_1y  : {date_str: {code: %}}     IMM日付開始 1Y fwd

    QuantLib はプロセス内グローバル状態を持つため _ql_lock で直列化。
    """
    carry_results:  dict = {}
    inst_results:   dict = {}
    imm_results_3m: dict = {}
    imm_results_1y: dict = {}

    with _ql_lock:
        for date_str, ois_rates in sorted(dates_ois.items()):
            curve_input = [
                {"tenor": t, "rate": r}
                for t, r in ois_rates.items()
                if r is not None
            ]
            if len(curve_input) < 4:
                continue

            try:
                ql_helper = QuantLibHelper(date_str)
                ql_helper.build_ois_curve(curve_input)
            except Exception as e:
                logger.debug(f"OIS curve build failed for {date_str}: {e}")
                continue

            carry_row: dict = {}
            inst_row:  dict = {}
            imm_row_3m: dict = {}
            imm_row_1y: dict = {}

            for tenor in carry_tenors:
                spot = ois_rates.get(tenor)
                if spot is None:
                    continue

                months = _tenor_to_months(tenor)
                fwd_months = months - 3
                if fwd_months <= 0:
                    continue
                fwd_tenor_str = f"{int(round(fwd_months))}M"

                # Carry-roll: 3M×(N-3M) fwd − spot(N)  [bps]
                try:
                    cr_rate = ql_helper.calculate_forward_swap_rate("3M", fwd_tenor_str)
                    if cr_rate is not None:
                        carry_row[tenor] = round((cr_rate - spot) * 100, 2)
                except Exception as e:
                    logger.debug(f"carry-roll failed {date_str} {tenor}: {e}")

                # Inst fwd: (N-3M)×3M fwd  [%]
                try:
                    inst_rate = ql_helper.calculate_forward_swap_rate(fwd_tenor_str, "3M")
                    if inst_rate is not None:
                        inst_row[tenor] = round(inst_rate, 4)
                except Exception as e:
                    logger.debug(f"inst-fwd failed {date_str} {tenor}: {e}")

            # IMM strip: 3M fwd / 1Y fwd
            for code, imm_date_str in imm_cols:
                try:
                    r3m = ql_helper.calculate_forward_from_start_date(imm_date_str, "3M")
                    if r3m is not None:
                        imm_row_3m[code] = round(r3m, 4)
                except Exception as e:
                    logger.debug(f"IMM 3M failed {date_str} {code}: {e}")

                try:
                    r1y = ql_helper.calculate_forward_from_start_date(imm_date_str, "1Y")
                    if r1y is not None:
                        imm_row_1y[code] = round(r1y, 4)
                except Exception as e:
                    logger.debug(f"IMM 1Y failed {date_str} {code}: {e}")

            carry_results[date_str]  = carry_row
            inst_results[date_str]   = inst_row
            imm_results_3m[date_str] = imm_row_3m
            imm_results_1y[date_str] = imm_row_1y

    return carry_results, inst_results, imm_results_3m, imm_results_1y


# ─────────────────────────────────────────────
# エンドポイント
# ─────────────────────────────────────────────
@router.get(
    "/asw-matrix",
    summary="ASW / TONA / フォワード 統合 Excel ダウンロード",
    response_class=StreamingResponse,
)
async def download_asw_matrix(
    lookback: int = Query(DEFAULT_LOOKBACK_DAYS, ge=30, le=1000,
                          description="過去何日分を取得するか（暦日）"),
):
    """
    5シート構成の Excel を返す。

    - ASW_SASA_5Y   : 残存5年以下クーポン国債の SASA ASW [bps]
    - TONA_HIST_1Y  : TONA OIS スワップ金利履歴 [%]
    - TONA_FWD_3M   : キャリーロール 3M×(N-3M) fwd − Spot(N) [bps]
    - TONA_INST_FWD : 瞬間フォワード (N-3M)×3M fwd [%]
    - TONA_IMM      : IMMストリップ 各IMM日付開始3M fwd [%]

    共通: 行1-2=ヘッダー、行3-7=Z-score数式、行8-12=空白、行13-=データ(降順)
    """
    start_date = _jst_today() - timedelta(days=lookback)
    today = _jst_today()

    try:
        # ── 1. ASW 銘柄ユニバース ───────────────────────────────
        async with AsyncSessionLocal() as session:
            rows = await session.execute(text("""
                SELECT b.bond_code, b.bond_name, b.due_date::text AS due_date
                FROM bond_data b
                JOIN "ASW_data" a
                  ON a.bond_code = b.bond_code
                 AND b.trade_date = (SELECT MAX(trade_date) FROM bond_data)
                WHERE b.trade_date = (SELECT MAX(trade_date) FROM bond_data)
                  AND b.due_date >  CURRENT_DATE
                  AND b.due_date <= CURRENT_DATE + make_interval(years => :maturity_years)
                GROUP BY b.bond_code, b.bond_name, b.due_date
                ORDER BY b.due_date, b.bond_code
            """), {"maturity_years": MAX_MATURITY_YEARS})
            universe = rows.fetchall()

        if not universe:
            raise HTTPException(status_code=404, detail="No bonds found.")

        bond_codes = [r.bond_code for r in universe]
        due_dates  = {r.bond_code: r.due_date for r in universe}

        # ── 2. ASW 履歴 ─────────────────────────────────────────
        async with AsyncSessionLocal() as session:
            rows = await session.execute(text("""
                SELECT a.trade_date::text AS trade_date, a.bond_code,
                       ROUND(a.asw_act365_sa::numeric * 100, 2) AS asw_bps
                FROM "ASW_data" a
                WHERE a.bond_code = ANY(:codes)
                  AND a.trade_date >= :start
                ORDER BY a.trade_date DESC, a.bond_code
            """), {"codes": bond_codes, "start": start_date})
            asw_history = rows.fetchall()

        # ── 3. TONA 履歴 ─────────────────────────────────────────
        async with AsyncSessionLocal() as session:
            rows = await session.execute(text("""
                SELECT trade_date::text AS trade_date, tenor,
                       ROUND(rate::numeric, 4) AS rate
                FROM irs_data
                WHERE product_type = 'OIS'
                  AND tenor = ANY(:tenors)
                  AND trade_date >= :start
                ORDER BY trade_date DESC, tenor
            """), {"tenors": STANDARD_OIS_TENORS, "start": start_date})
            tona_history = rows.fetchall()

        if not tona_history:
            raise HTTPException(status_code=404, detail="No TONA data found.")

        # ── 4. データ集約 ────────────────────────────────────────
        from collections import defaultdict

        asw_pivot: dict = defaultdict(dict)
        for r in asw_history:
            asw_pivot[r.trade_date][r.bond_code] = r.asw_bps
        asw_rows = sorted(asw_pivot.items(), reverse=True)

        tona_pivot: dict = defaultdict(dict)
        for r in tona_history:
            tona_pivot[r.trade_date][r.tenor] = float(r.rate)
        tona_rows_desc = sorted(tona_pivot.items(), reverse=True)

        carry_tenors = [t for t in STANDARD_OIS_TENORS if _tenor_to_months(t) > 3]

        # ── 5. IMM 列生成 ────────────────────────────────────────
        imm_cols = _get_imm_strip_columns(today, years=10)
        logger.info(f"IMM strip: {len(imm_cols)} columns ({imm_cols[0][0]}..{imm_cols[-1][0]})")

        # ── 6. QuantLib 全フォワードレート計算 ───────────────────
        logger.info(
            f"Starting forward rate computation: "
            f"{len(tona_pivot)} dates × {len(carry_tenors)} tenors + {len(imm_cols)} IMM cols"
        )
        carry_results, inst_results, imm_results_3m, imm_results_1y = await run_in_threadpool(
            _compute_all_forward_rates_sync,
            dict(tona_pivot),
            carry_tenors,
            imm_cols,
        )
        carry_rows_desc    = sorted(carry_results.items(),  reverse=True)
        inst_rows_desc     = sorted(inst_results.items(),   reverse=True)
        imm_rows_3m_desc   = sorted(imm_results_3m.items(), reverse=True)
        imm_rows_1y_desc   = sorted(imm_results_1y.items(), reverse=True)
        logger.info(f"Forward rate computation done: {len(carry_results)} dates")

        # ── 7. Excel 生成 ────────────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: ASW
        _build_asw_sheet(wb, bond_codes, due_dates, asw_rows, "ASW_SASA_5Y")

        # Sheet 2: TONA History
        available_tona_tenors = [
            t for t in STANDARD_OIS_TENORS
            if any(t in d for _, d in tona_rows_desc)
        ]
        _build_tona_hist_sheet(wb, available_tona_tenors, tona_rows_desc, "TONA_HIST_1Y")

        # Sheet 3: Carry-Roll
        available_carry_tenors = [
            t for t in carry_tenors
            if any(t in d for _, d in carry_rows_desc)
        ]
        _build_carry_roll_sheet(wb, available_carry_tenors, carry_rows_desc, "TONA_FWD_3M")

        # Sheet 4: Instantaneous Forward
        available_inst_tenors = [
            t for t in carry_tenors
            if any(t in d for _, d in inst_rows_desc)
        ]
        _build_tona_inst_fwd_sheet(wb, available_inst_tenors, inst_rows_desc, "TONA_INST_FWD")

        # Sheet 5: IMM Strip 3M
        available_imm_cols_3m = [
            (code, ds) for code, ds in imm_cols
            if any(code in d for _, d in imm_rows_3m_desc)
        ]
        _build_imm_strip_sheet(wb, available_imm_cols_3m, imm_rows_3m_desc, "TONA_IMM_3M")

        # Sheet 6: IMM Strip 1Y
        available_imm_cols_1y = [
            (code, ds) for code, ds in imm_cols
            if any(code in d for _, d in imm_rows_1y_desc)
        ]
        _build_imm_strip_sheet(wb, available_imm_cols_1y, imm_rows_1y_desc, "TONA_IMM_1Y")

        # ── 8. ストリーム返却 ─────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)  # getvalue() + new BytesIO の二重確保を回避

        filename = f"market_data_{today.strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel export error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────
# IMM専用計算（軽量版: carry/inst は計算しない）
# ─────────────────────────────────────────────
def _compute_imm_spot_rates_sync(
    dates_ois: dict,   # {date_str: {tenor: rate_pct}}
    imm_cols: list,    # [(code, date_str), ...]
) -> dict:             # {date_str: {code: rate_%}}
    """
    各営業日について OISカーブを構築し、
    スポット日（eval_date + 2営業日）から各IMM日付までのOISパースワップレートを計算する。
    例: Z31 → today+2 から 2031-12-17 までの OIS スワップレート [%]
    """
    imm_results: dict = {}

    with _ql_lock:
        for date_str, ois_rates in sorted(dates_ois.items()):
            curve_input = [
                {"tenor": t, "rate": r}
                for t, r in ois_rates.items()
                if r is not None
            ]
            if len(curve_input) < 4:
                continue

            try:
                ql_helper = QuantLibHelper(date_str)
                ql_helper.build_ois_curve(curve_input)
            except Exception as e:
                logger.debug(f"OIS curve build failed for {date_str}: {e}")
                continue

            row: dict = {}
            for code, imm_date_str in imm_cols:
                try:
                    r = ql_helper.calculate_spot_ois_to_date(imm_date_str)
                    if r is not None:
                        row[code] = r
                except Exception as e:
                    logger.debug(f"IMM spot OIS failed {date_str} {code}: {e}")

            imm_results[date_str] = row

    return imm_results


# ─────────────────────────────────────────────
# IMM Rates 専用 Excel エンドポイント
# ─────────────────────────────────────────────
@router.get(
    "/imm-rates",
    summary="IMM OIS フォワードレート Excel ダウンロード（~160列）",
    response_class=StreamingResponse,
)
async def download_imm_rates(
    lookback: int = Query(200, ge=30, le=1000,
                          description="過去何日分を取得するか（暦日、デフォルト200）"),
):
    """
    過去N日分の IMM ストリップ OIS フォワードレートを Excel で返す。

    - IMM_3M : 各IMM日付開始の 3M OIS フォワードレート [%]
    - IMM_1Y : 各IMM日付開始の 1Y OIS フォワードレート [%]

    列構成:
      行1 : IMMコード（U26, Z26, H27, ...  ~160列、40年分）
      行2 : 実際のIMM日付（第3水曜日）
      行3-7 : Z-score（10D/20D/50D/100D/200D）Excel数式
      行8-12: 空白
      行13- : データ（新しい日付から降順）
    """
    start_date = _jst_today() - timedelta(days=lookback)
    today = _jst_today()

    try:
        # ── 1. TONA 履歴取得 ──────────────────────────────────────
        async with AsyncSessionLocal() as session:
            rows = await session.execute(text("""
                SELECT trade_date::text AS trade_date, tenor,
                       ROUND(rate::numeric, 4) AS rate
                FROM irs_data
                WHERE product_type = 'OIS'
                  AND tenor = ANY(:tenors)
                  AND trade_date >= :start
                ORDER BY trade_date DESC, tenor
            """), {"tenors": STANDARD_OIS_TENORS, "start": start_date})
            tona_history = rows.fetchall()

        if not tona_history:
            raise HTTPException(status_code=404, detail="No TONA/OIS data found.")

        # ── 2. データ集約 ──────────────────────────────────────────
        from collections import defaultdict

        tona_pivot: dict = defaultdict(dict)
        for r in tona_history:
            tona_pivot[r.trade_date][r.tenor] = float(r.rate)

        # ── 3. IMM 列生成（40年分 ≈ 160列）──────────────────────────
        imm_cols = _get_imm_strip_columns(today, years=40, max_cols=160)
        logger.info(
            f"IMM-only export: {len(imm_cols)} columns "
            f"({imm_cols[0][0]}..{imm_cols[-1][0]}), "
            f"{len(tona_pivot)} dates"
        )

        # ── 4. QuantLib IMM スポットレート計算 ────────────────────
        imm_results = await run_in_threadpool(
            _compute_imm_spot_rates_sync,
            dict(tona_pivot),
            imm_cols,
        )
        imm_rows_desc = sorted(imm_results.items(), reverse=True)
        logger.info(f"IMM spot computation done: {len(imm_results)} dates")

        # ── 5. Excel 生成 ──────────────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)

        # 実際にデータがある列のみ抽出
        available_cols = [
            (code, ds) for code, ds in imm_cols
            if any(code in d for _, d in imm_rows_desc)
        ]

        _build_imm_standalone_sheet(wb, available_cols, imm_rows_desc, "IMM_SPOT_OIS", reference_date=today)

        # ── 6. ストリーム返却 ──────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"imm_rates_{today.strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"IMM Excel export error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
