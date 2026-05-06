#!/usr/bin/env python3
"""
ASW SASA マトリックス生成スクリプト
====================================
現在残存5年以下の全クーポン国債について、過去1年分のASW（SASA方式）を
マトリックス形式のExcelシートに出力する。

使い方:
    ./venv/bin/python scripts/export/generate_asw_matrix.py
    ./venv/bin/python scripts/export/generate_asw_matrix.py --output exports/my_file.xlsx
    ./venv/bin/python scripts/export/generate_asw_matrix.py --sheet "ASW_SASA_5Y"

出力形式:
    行1  : bond_code（列ヘッダー）
    行2  : due_date（参照用）
    行3- : trade_date + ASW値（bps、小数第2位まで）

ASW計算方式: SASA（Semi-Annual固定脚 / Act365 vs TONA OIS）
単位: bps（データはDB内で%保存 → 出力時に×100してbps変換）
"""

import argparse
import sys
import os
from datetime import date, timedelta
from pathlib import Path

# プロジェクトルートをPATHに追加
project_root = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(project_root))

import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.config import settings
from core.db.sync_client import DatabaseManager


# ─────────────────────────────────────────────
# 定数
# ─────────────────────────────────────────────
DEFAULT_OUTPUT_DIR = project_root / "exports"
LOOKBACK_DAYS = 365          # 過去何日分を対象とするか（暦日）
MAX_MATURITY_YEARS = 5       # 残存年数の上限

# Excel スタイル
HEADER1_FILL = PatternFill("solid", fgColor="1F4E79")   # 濃紺 - bond_code行
HEADER1_FONT = Font(color="FFFFFF", bold=True, size=9)
HEADER2_FILL = PatternFill("solid", fgColor="2E75B6")   # 中青 - due_date行
HEADER2_FONT = Font(color="FFFFFF", bold=True, size=8)
DATE_FILL    = PatternFill("solid", fgColor="D6E4F0")   # 薄青 - trade_date列
DATE_FONT    = Font(bold=True, size=9)
DATA_FONT    = Font(size=9)
THIN_BORDER  = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _get_engine():
    """SQLAlchemy同期エンジンを作成（pandas互換）"""
    return create_engine(settings.database_url)


def fetch_bond_universe(engine) -> pd.DataFrame:
    """
    最新取引日時点で残存5年以下かつASWデータが存在するクーポン銘柄一覧を取得。
    Returns: DataFrame [bond_code, bond_name, due_date] (due_date昇順)
    """
    sql = text(f"""
        SELECT DISTINCT b.bond_code, b.bond_name, b.due_date
        FROM bond_data b
        JOIN "ASW_data" a
          ON a.bond_code = b.bond_code
         AND b.trade_date = (SELECT MAX(trade_date) FROM bond_data)
        WHERE b.trade_date = (SELECT MAX(trade_date) FROM bond_data)
          AND b.due_date >  CURRENT_DATE
          AND b.due_date <= CURRENT_DATE + INTERVAL '{MAX_MATURITY_YEARS} years'
        ORDER BY b.due_date, b.bond_code
    """)
    with engine.connect() as conn:
        return pd.read_sql(sql, conn)


def fetch_asw_history(engine, bond_codes: list, start_date: date) -> pd.DataFrame:
    """
    指定銘柄の過去ASW（SASA = asw_act365_sa）を取得。
    Returns: DataFrame [trade_date, bond_code, asw_bps]
    """
    sql = text("""
        SELECT a.trade_date, a.bond_code,
               ROUND(a.asw_act365_sa::numeric * 100, 2) AS asw_bps
        FROM "ASW_data" a
        WHERE a.bond_code = ANY(:codes)
          AND a.trade_date >= :start_date
        ORDER BY a.trade_date, a.bond_code
    """)
    with engine.connect() as conn:
        return pd.read_sql(sql, conn, params={"codes": bond_codes, "start_date": start_date})


def build_matrix(history: pd.DataFrame, universe: pd.DataFrame) -> pd.DataFrame:
    """
    (trade_date × bond_code) のピボットマトリックスを生成。
    列順は due_date 昇順（universe の順序を維持）。
    """
    pivot = history.pivot(index="trade_date", columns="bond_code", values="asw_bps")

    # 列を due_date 昇順に並び替え（universe の bond_code 順序を使用）
    ordered_codes = [c for c in universe["bond_code"] if c in pivot.columns]
    pivot = pivot[ordered_codes]

    # 行を昇順（古い→新しい）に並び替え
    pivot = pivot.sort_index(ascending=True)
    pivot.index = pd.to_datetime(pivot.index).date  # date型に統一

    return pivot


def write_excel(
    matrix: pd.DataFrame,
    universe: pd.DataFrame,
    output_path: Path,
    sheet_name: str,
) -> None:
    """
    マトリックスをExcelに書き出す。
    既存ファイルがあればシートを更新、なければ新規作成。
    """
    from openpyxl import load_workbook

    # 既存ファイルがあれば読み込み、シートのみ差し替え
    if output_path.exists():
        wb = load_workbook(output_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    bond_codes = list(matrix.columns)
    due_date_map = universe.set_index("bond_code")["due_date"].to_dict()

    # ── 行1: bond_code ヘッダー ──────────────────────────────
    ws.cell(1, 1, "trade_date").font = HEADER1_FONT
    ws.cell(1, 1).fill = HEADER1_FILL
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(1, 1).border = THIN_BORDER

    for col_idx, code in enumerate(bond_codes, start=2):
        cell = ws.cell(1, col_idx, code)
        cell.font = HEADER1_FONT
        cell.fill = HEADER1_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # ── 行2: due_date ヘッダー ────────────────────────────────
    ws.cell(2, 1, "due_date").font = HEADER2_FONT
    ws.cell(2, 1).fill = HEADER2_FILL
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(2, 1).border = THIN_BORDER

    for col_idx, code in enumerate(bond_codes, start=2):
        dd = due_date_map.get(code, "")
        cell = ws.cell(2, col_idx, str(dd) if dd else "")
        cell.font = HEADER2_FONT
        cell.fill = HEADER2_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # ── 行3以降: データ ───────────────────────────────────────
    for row_idx, (trade_date, row_data) in enumerate(matrix.iterrows(), start=3):
        # trade_date 列
        date_cell = ws.cell(row_idx, 1, str(trade_date))
        date_cell.font = DATE_FONT
        date_cell.fill = DATE_FILL
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.border = THIN_BORDER

        # ASW値列
        for col_idx, code in enumerate(bond_codes, start=2):
            val = row_data.get(code)
            cell = ws.cell(row_idx, col_idx)
            if pd.notna(val):
                cell.value = float(val)
                cell.number_format = "0.00"
            else:
                cell.value = None
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = THIN_BORDER

    # ── 列幅の調整 ────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12  # trade_date列
    for col_idx in range(2, len(bond_codes) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11

    # ── 行高の調整 ────────────────────────────────────────────
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16

    # ── ウィンドウ枠固定（行1-2、列A を固定） ─────────────────
    ws.freeze_panes = "B3"

    wb.save(output_path)


def parse_args():
    today_str = date.today().strftime("%Y%m%d")
    parser = argparse.ArgumentParser(
        description="ASW SASA マトリックスをExcelに出力する"
    )
    parser.add_argument(
        "--output", "-o",
        default=str(DEFAULT_OUTPUT_DIR / f"asw_matrix_{today_str}.xlsx"),
        help="出力Excelファイルパス (default: exports/asw_matrix_YYYYMMDD.xlsx)",
    )
    parser.add_argument(
        "--sheet", "-s",
        default="ASW_SASA_5Y",
        help="シート名 (default: ASW_SASA_5Y)",
    )
    parser.add_argument(
        "--lookback", "-l",
        type=int,
        default=LOOKBACK_DAYS,
        help=f"過去何日分を取得するか（暦日, default: {LOOKBACK_DAYS}）",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    start_date = date.today() - timedelta(days=args.lookback)

    print(f"=== ASW SASA マトリックス生成 ===")
    print(f"  対象残存年数   : {MAX_MATURITY_YEARS}年以下")
    print(f"  データ期間     : {start_date} 〜 今日")
    print(f"  ASW方式        : SASA (Semi-Annual / Act365 vs TONA OIS)")
    print(f"  出力ファイル   : {output_path}")
    print(f"  シート名       : {args.sheet}")
    print()

    engine = _get_engine()

    # 1. 銘柄ユニバース取得
    print("1/4 銘柄ユニバース取得中...")
    universe = fetch_bond_universe(engine)
    print(f"    → {len(universe)} 銘柄")

    if universe.empty:
        print("ERROR: 銘柄が見つかりませんでした。")
        sys.exit(1)

    # 2. ASW履歴取得
    print("2/4 ASW履歴取得中...")
    history = fetch_asw_history(engine, universe["bond_code"].tolist(), start_date)
    print(f"    → {len(history)} レコード / {history['trade_date'].nunique()} 営業日")

    if history.empty:
        print("ERROR: ASWデータが見つかりませんでした。")
        sys.exit(1)

    # 3. マトリックス構築
    print("3/4 マトリックス構築中...")
    matrix = build_matrix(history, universe)
    print(f"    → {matrix.shape[0]} 行 × {matrix.shape[1]} 列")

    # 4. Excel出力
    print("4/4 Excel出力中...")
    write_excel(matrix, universe, output_path, args.sheet)
    print(f"    → 完了: {output_path}")

    print()
    print("=== 完了 ===")
    print(f"  営業日数 : {matrix.shape[0]}")
    print(f"  銘柄数   : {matrix.shape[1]}")
    print(f"  単位     : bps（小数第2位まで）")
    print(f"  ファイル : {output_path.resolve()}")


if __name__ == "__main__":
    main()
